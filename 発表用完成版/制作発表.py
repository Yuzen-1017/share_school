import tkinter as tk
from tkinter import messagebox, filedialog
import os
import configparser
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime

# 設定ファイルとログファイルのパス
config_file = 'config.ini'
log_file = 'report_logs.txt'

# セル位置設定ウィンドウ
cell_setting_window = None

# セル位置設定ウィンドウ内のエントリ変数
entry_name_cell = None
entry_date_cell = None
entry_content_cell = None

# テンプレートファイルを選択する関数
def select_template():
    file_path = filedialog.askopenfilename(filetypes=[("Excel ファイル", "*.xlsx")])
    entry_template.delete(0, tk.END)
    entry_template.insert(tk.END, file_path)
    update_config(file_path)

# 設定ファイルを更新する関数
def update_config(template_file):
    config = configparser.ConfigParser()
    config.read(config_file)
    if not config.has_section('TemplateInfo'):
        config.add_section('TemplateInfo')

    filename, _ = os.path.splitext(os.path.basename(template_file))
    if filename not in config.options('TemplateInfo'):
        config.set('TemplateInfo', filename, '')
        with open(config_file, 'w') as configfile:
            config.write(configfile)

# セル位置設定ウィンドウを表示する関数
def set_cell_positions():
    global cell_setting_window, entry_name_cell, entry_date_cell, entry_content_cell

    template_file = entry_template.get()
    if not template_file:
        messagebox.showerror('エラー', 'テンプレートファイルを選択してください。')
        return

    if not os.path.exists(template_file):
        messagebox.showerror('エラー', '指定されたテンプレートファイルが存在しません。')
        return

    cell_setting_window = tk.Toplevel(root)
    cell_setting_window.title('セル位置設定')

    label_instruction = tk.Label(cell_setting_window, text='以下の各項目について、該当するセルを選択してください:')
    label_instruction.grid(row=0, column=0, columnspan=2, pady=5, sticky='w')

    label_name_cell = tk.Label(cell_setting_window, text='名前:')
    label_name_cell.grid(row=1, column=0, pady=5, sticky='w')

    label_date_cell = tk.Label(cell_setting_window, text='日付:')
    label_date_cell.grid(row=2, column=0, pady=5, sticky='w')

    label_content_cell = tk.Label(cell_setting_window, text='内容:')
    label_content_cell.grid(row=3, column=0, pady=5, sticky='w')

    entry_name_cell = tk.Entry(cell_setting_window)
    entry_name_cell.grid(row=1, column=1, pady=5, sticky='w')

    entry_date_cell = tk.Entry(cell_setting_window)
    entry_date_cell.grid(row=2, column=1, pady=5, sticky='w')

    entry_content_cell = tk.Entry(cell_setting_window)
    entry_content_cell.grid(row=3, column=1, pady=5, sticky='w')

    button_save = tk.Button(cell_setting_window, text='保存', command=save_cell_positions)
    button_save.grid(row=4, columnspan=2, pady=10)

# セル位置を保存する関数
def save_cell_positions():
    global cell_setting_window, entry_name_cell, entry_date_cell, entry_content_cell

    name_cell = entry_name_cell.get()
    date_cell = entry_date_cell.get()
    content_cell = entry_content_cell.get()

    if not (name_cell and date_cell and content_cell):
        messagebox.showerror('エラー', '全てのセルを指定してください。')
        return

    config = configparser.ConfigParser()
    config.read(config_file)

    template_file = entry_template.get()
    filename, _ = os.path.splitext(os.path.basename(template_file))
    config.set('TemplateInfo', filename, f'{name_cell},{date_cell},{content_cell}')

    with open(config_file, 'w') as configfile:
        config.write(configfile)

    cell_setting_window.destroy()

# 日報を挿入する関数
def insert_report():
    template_file = entry_template.get()
    name = entry_name.get().strip()
    date = entry_date.get().strip()
    content = text_content.get('1.0', 'end-1c').strip()

    if not (template_file and name and date and content):
        messagebox.showerror('エラー', '全ての項目を入力してください。')
        return

    config = configparser.ConfigParser()
    config.read(config_file)
    filename, _ = os.path.splitext(os.path.basename(template_file))

    if not config.has_option('TemplateInfo', filename):
        messagebox.showerror('エラー', 'テンプレートファイルのセル位置が設定されていません。')
        return

    name_cell, date_cell, content_cell = config.get('TemplateInfo', filename).split(',')

    try:
        wb = load_workbook(template_file)
        ws = wb.active

        # 日付を入力
        ws[name_cell] = name
        ws[date_cell] = date 
        ws[content_cell] = content
        ws[content_cell].alignment = Alignment(wrap_text=True, vertical='top')
        
        # 新しいファイルを作成して保存
        filename, extension = os.path.splitext(template_file)
        modified_file = f"{filename}_modified{extension}"
        wb.save(modified_file)

        with open(log_file, 'a') as logfile:
            logfile.write(f"{datetime.now()} 日報が入力されました。日付: {date}, 名前: {name}, 内容: {content}\n")

        messagebox.showinfo('成功', '日報が保存されました。')

    except Exception as e:
        messagebox.showerror('エラー', f'エラーが発生しました: {str(e)}')

# GUIの構築
root = tk.Tk()
root.title('日報入力画面')

label_template = tk.Label(root, text='テンプレートファイル:')
label_template.grid(row=0, column=0, sticky='w')
entry_template = tk.Entry(root, width=50)
entry_template.grid(row=0, column=1, pady=5, sticky='w')
button_browse = tk.Button(root, text='参照', command=select_template)
button_browse.grid(row=0, column=2, padx=5)

button_set_positions = tk.Button(root, text='セル位置設定', command=set_cell_positions)
button_set_positions.grid(row=1, columnspan=2, pady=5)

label_name = tk.Label(root, text='名前:')
label_name.grid(row=2, column=0, sticky='w')
entry_name = tk.Entry(root)
entry_name.grid(row=2, column=1, pady=5, sticky='w')

label_date = tk.Label(root, text='日付(YYYY年MM月DD日):')
label_date.grid(row=3, column=0, sticky='w')
entry_date = tk.Entry(root)
entry_date.grid(row=3, column=1, pady=5, sticky='w')

label_content = tk.Label(root, text='内容:')
label_content.grid(row=4, column=0, sticky="w")
text_content = tk.Text(root, height=10, width=50)
text_content.grid(row=4, column=1, columnspan=2, pady=5, sticky="w")

button_save = tk.Button(root, text='保存', command=insert_report)
button_save.grid(row=5, columnspan=2, pady=10)

root.mainloop()
